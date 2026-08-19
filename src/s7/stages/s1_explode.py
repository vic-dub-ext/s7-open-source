"""S1 - Explode. In: artifacts from S0. Out: one artifact per worksheet for
every workbook, plus a wrapped single-sheet artifact for each CSV/TSV. PDFs
and everything else pass through unchanged.

This stage exists because Extend's classifier assigns one category per
document -- a 30-sheet workbook classified whole returns one useless label.
"""

from __future__ import annotations

import csv
import io
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy import Connection

from s7.config import Settings, get_settings
from s7.models.stage import StageResult, StageStatus
from s7.storage import store_bytes
from s7.store import events
from s7.store.artifacts import (
    delete_children_for_run,
    has_children,
    insert_artifact,
    list_top_level,
)
from s7.store.db import get_engine, now_iso
from s7.store.runs import update_run_status

STAGE = "s1_explode"

WORKBOOK_EXTENSIONS = {".xlsx", ".xls"}
DELIMITED_EXTENSIONS = {".csv": ",", ".tsv": "\t"}
MIN_ROWS = 2
MIN_COLS = 2
CLASSIFY_SAMPLE_ROW_CAP = 5000
CLASSIFY_SAMPLE_HEADER_ROWS = 1
CLASSIFY_SAMPLE_DATA_ROWS = 200


def _ext(file_name: str) -> str:
    return Path(file_name).suffix.lower()


def _explode_workbook(
    conn: Connection, *, run_id: str, parent: dict[str, Any], settings: Settings
) -> tuple[int, int]:
    """Split every worksheet in `parent` into its own single-sheet artifact.
    Returns (sheets_written, empty_sheets).
    """
    storage_path = Path(str(parent["storage_path"]))
    data = storage_path.read_bytes()

    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except (InvalidFileException, KeyError, OSError) as exc:
        events.emit(
            conn,
            run_id=run_id,
            stage=STAGE,
            event_type="error",
            level="warn",
            message=f"could not open {parent['file_name']} as a workbook: {exc}",
        )
        return 0, 0

    written = 0
    empty = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        row_count = len(rows)
        col_count = max((len(r) for r in rows), default=0)
        is_empty = row_count < MIN_ROWS or col_count < MIN_COLS

        cap = CLASSIFY_SAMPLE_HEADER_ROWS + CLASSIFY_SAMPLE_DATA_ROWS
        is_sample = row_count > CLASSIFY_SAMPLE_ROW_CAP
        sheet_rows = rows[:cap] if is_sample else rows

        out_wb = openpyxl.Workbook()
        out_ws = out_wb.active
        assert out_ws is not None
        out_ws.title = sheet_name[:31] or "Sheet1"
        for row in sheet_rows:
            out_ws.append(list(row))
        buf = io.BytesIO()
        out_wb.save(buf)

        digest, path = store_bytes(settings.downloads_dir, buf.getvalue())
        insert_artifact(
            conn,
            run_id=run_id,
            kind="sheet",
            file_name=f"{parent['file_name']}::{sheet_name}",
            mime_type=str(parent["mime_type"]),
            byte_size=buf.getbuffer().nbytes,
            sha256=digest,
            storage_path=str(path),
            retrieved_at=now_iso(),
            parent_artifact_id=str(parent["id"]),
            sheet_name=sheet_name,
            row_offset=0,  # the split file mirrors the original sheet 1:1, no cropping
            col_offset=0,
            row_count=row_count,
            col_count=col_count,
            is_classify_sample=is_sample,
            skip_reason="empty_sheet" if is_empty else None,
            skip_detail="fewer than 2 rows or 2 columns" if is_empty else None,
        )
        written += 1
        empty += int(is_empty)

    wb.close()
    return written, empty


def _wrap_delimited(
    conn: Connection, *, run_id: str, parent: dict[str, Any], delimiter: str
) -> int:
    """CSV/TSV files are already one flat table -- wrap as a single sheet
    artifact pointing at the same bytes, for uniformity with exploded sheets.
    """
    storage_path = Path(str(parent["storage_path"]))
    text = storage_path.read_text(encoding="utf-8", errors="replace")
    rows = list(csv.reader(io.StringIO(text), delimiter=delimiter))
    row_count = len(rows)
    col_count = max((len(r) for r in rows), default=0)
    is_empty = row_count < MIN_ROWS or col_count < MIN_COLS

    insert_artifact(
        conn,
        run_id=run_id,
        kind="sheet",
        file_name=str(parent["file_name"]),
        mime_type=str(parent["mime_type"]),
        byte_size=int(parent["byte_size"]),
        sha256=str(parent["sha256"]),
        storage_path=str(parent["storage_path"]),
        retrieved_at=now_iso(),
        parent_artifact_id=str(parent["id"]),
        sheet_name=None,
        row_offset=0,
        col_offset=0,
        row_count=row_count,
        col_count=col_count,
        is_classify_sample=False,
        skip_reason="empty_sheet" if is_empty else None,
        skip_detail="fewer than 2 rows or 2 columns" if is_empty else None,
    )
    return 1


async def run(run_id: str, *, force: bool = False) -> StageResult:
    settings = get_settings()
    settings.ensure_dirs()
    engine = get_engine()

    with engine.begin() as conn:
        already_exploded = has_children(conn, run_id)
        if not force and already_exploded:
            return StageResult(stage=STAGE, status="skipped", counts={})
        if force and already_exploded:
            delete_children_for_run(conn, run_id)

        events.emit(
            conn, run_id=run_id, stage=STAGE, event_type="stage_started", message="S1 started"
        )
        update_run_status(conn, run_id, status="running", stage_reached=STAGE)
        parents = [a for a in list_top_level(conn, run_id) if not a["skip_reason"]]

        sheets_written = 0
        empty_sheets = 0
        passthrough = 0

        for parent in parents:
            ext = _ext(str(parent["file_name"]))
            if ext in WORKBOOK_EXTENSIONS:
                written, empty = _explode_workbook(
                    conn, run_id=run_id, parent=parent, settings=settings
                )
                sheets_written += written
                empty_sheets += empty
            elif ext in DELIMITED_EXTENSIONS:
                sheets_written += _wrap_delimited(
                    conn, run_id=run_id, parent=parent, delimiter=DELIMITED_EXTENSIONS[ext]
                )
            else:
                passthrough += 1  # PDFs and anything else pass through unchanged

        status: StageStatus = "done"
        update_run_status(conn, run_id, status=status, stage_reached=STAGE)
        events.emit(
            conn,
            run_id=run_id,
            stage=STAGE,
            event_type="stage_finished",
            message=f"S1 explode finished: {sheets_written} sheets "
            f"({empty_sheets} empty), {passthrough} passed through unchanged",
            payload={
                "sheets": sheets_written,
                "empty": empty_sheets,
                "passthrough": passthrough,
            },
        )

    return StageResult(
        stage=STAGE,
        status=status,
        counts={"sheets": sheets_written, "empty": empty_sheets, "passthrough": passthrough},
    )
